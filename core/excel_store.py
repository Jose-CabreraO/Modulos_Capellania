from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


BASE_DIR = Path(__file__).resolve().parents[1]
EXCEL_FILE = BASE_DIR / "plantilla_flujo_completo.xlsx"
SHEET_GRUPOS = "1_Crear_Grupos"
SHEET_INTEGRANTES = "2_Agregar_Integrantes"
SHEET_REFERENCIAS = "Listas_Referencia"
STATUS_COL = "Estado_Carga"

FILL_PROCESADO = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
FILL_ERROR = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
FILL_LIMPIO = PatternFill(fill_type=None)
GROUP_COLUMNS = [
    "Grupo / Nombre*",
    "Empresa*",
    "Sucursal*",
    "Capellán*",
    "Material*",
    "Estado Del Grupo*",
    STATUS_COL,
]


def inicializar_estado_carga(excel_file=EXCEL_FILE):
    wb = load_workbook(excel_file)
    try:
        ws = wb[SHEET_GRUPOS]
        headers = [cell.value for cell in ws[1]]
        cambiado = False

        if STATUS_COL not in headers:
            status_col = ws.max_column + 1
            ws.cell(row=1, column=status_col, value=STATUS_COL)
            cambiado = True
        else:
            status_col = headers.index(STATUS_COL) + 1

        for row in range(2, ws.max_row + 1):
            nombre = ws.cell(row=row, column=1).value
            estado = ws.cell(row=row, column=status_col).value
            if nombre and not estado:
                ws.cell(row=row, column=status_col, value="PENDIENTE")
                cambiado = True

        if cambiado:
            wb.save(excel_file)
    finally:
        wb.close()


def leer_grupos(excel_file=EXCEL_FILE):
    inicializar_estado_carga(excel_file)
    return pd.read_excel(excel_file, sheet_name=SHEET_GRUPOS).dropna(subset=["Grupo / Nombre*"])


def leer_integrantes(excel_file=EXCEL_FILE):
    return pd.read_excel(excel_file, sheet_name=SHEET_INTEGRANTES).dropna(subset=["Nombre*"])


def leer_referencias(excel_file=EXCEL_FILE):
    try:
        df = pd.read_excel(excel_file, sheet_name=SHEET_REFERENCIAS)
    except Exception:
        return {}
    return {
        col: [str(v).strip() for v in df[col].dropna().tolist() if str(v).strip()]
        for col in df.columns
    }


def resumen_estados(excel_file=EXCEL_FILE):
    df = leer_grupos(excel_file)
    estados = df[STATUS_COL].fillna("PENDIENTE").astype(str).str.upper().str.strip()
    return {
        "PENDIENTE": int((estados == "PENDIENTE").sum()),
        "PROCESADO": int((estados == "PROCESADO").sum()),
        "ERROR": int((estados == "ERROR").sum()),
    }


def actualizar_estado_fila(excel_row, estado, fill, excel_file=EXCEL_FILE):
    wb = load_workbook(excel_file)
    try:
        ws = wb[SHEET_GRUPOS]
        status_col = 7
        ws.cell(row=1, column=status_col, value=STATUS_COL)

        ws.cell(row=excel_row, column=status_col, value=estado)
        for col in range(1, status_col + 1):
            ws.cell(row=excel_row, column=col).fill = fill

        wb.save(excel_file)
    finally:
        wb.close()


def _fill_por_estado(estado):
    estado = str(estado or "").strip().upper()
    if estado == "PROCESADO":
        return FILL_PROCESADO
    if estado == "ERROR":
        return FILL_ERROR
    return FILL_LIMPIO


def guardar_tabla_editada(df_editado, excel_file=EXCEL_FILE):
    df = df_editado.copy()
    for col in GROUP_COLUMNS:
        if col not in df.columns:
            df[col] = "PENDIENTE" if col == STATUS_COL else ""

    df = df[GROUP_COLUMNS].fillna("")
    df[STATUS_COL] = df[STATUS_COL].replace("", "PENDIENTE").astype(str).str.strip().str.upper()
    df["Estado Del Grupo*"] = df["Estado Del Grupo*"].replace("", "ACTIVO")
    df = df[df["Grupo / Nombre*"].astype(str).str.strip() != ""]

    wb = load_workbook(excel_file)
    try:
        ws = wb[SHEET_GRUPOS]

        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)

        for col_idx, header in enumerate(GROUP_COLUMNS, start=1):
            ws.cell(row=1, column=col_idx, value=header)

        for row_idx, row in enumerate(df.itertuples(index=False, name=None), start=2):
            fill = _fill_por_estado(row[GROUP_COLUMNS.index(STATUS_COL)])
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.fill = FILL_LIMPIO
                cell.value = value
                cell.fill = fill

        wb.save(excel_file)
    finally:
        wb.close()


def inyectar_whatsapp(registros, excel_file=EXCEL_FILE):
    inicializar_estado_carga(excel_file)
    wb = load_workbook(excel_file)
    try:
        ws_grupos = wb[SHEET_GRUPOS]
        ws_integrantes = wb[SHEET_INTEGRANTES]

        grupos_insertados = 0
        integrantes_insertados = 0
        for registro in registros:
            grupo = str(registro["Grupo / Nombre*"]).strip()
            ws_grupos.append([
                grupo,
                registro.get("Empresa*", ""),
                registro.get("Sucursal*", ""),
                registro.get("Capellán*", ""),
                registro.get("Material*", ""),
                registro.get("Estado Del Grupo*", "ACTIVO") or "ACTIVO",
                "PENDIENTE",
            ])
            grupos_insertados += 1

            for integrante in registro["Integrantes"]:
                ws_integrantes.append([
                    grupo,
                    integrante.get("Nombre*", ""),
                    integrante.get("Apellido", ""),
                    integrante.get("CI", ""),
                    integrante.get("Email", ""),
                    integrante.get("Sección", ""),
                ])
                integrantes_insertados += 1

        wb.save(excel_file)
        return grupos_insertados, integrantes_insertados
    finally:
        wb.close()
