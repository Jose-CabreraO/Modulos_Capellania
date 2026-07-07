import time
from urllib.parse import urljoin

import pandas as pd
from openpyxl.styles import PatternFill
from playwright.sync_api import sync_playwright

from core.config import capellania_credentials
from core.excel_store import (
    EXCEL_FILE,
    STATUS_COL,
    actualizar_estado_fila,
    inicializar_estado_carga,
    leer_grupos,
    leer_integrantes,
)


WP_LOGIN_URL = "https://capellania-app.visualweb.systems/wp-login.php"
TARGET_URL = "https://capellania-app.visualweb.systems/wp-admin/admin.php?page=capapp2_estudios_biblicos"
CHROMIUM_ARGS = ["--no-sandbox", "--disable-setuid-sandbox", "--disable-dev-shm-usage"]

green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
red_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")


def validar_credenciales():
    credentials = capellania_credentials()
    if not credentials["user"] or not credentials["password"]:
        raise RuntimeError(
            "Faltan CAPELLANIA_USER y/o CAPELLANIA_PASS. "
            "Definilas como variables de entorno antes de ejecutar la carga."
        )
    return credentials


def seleccionar_opcion_flexible(page, selector, texto_buscar, intentos=5):
    """
    Busca y selecciona una opcion ignorando acentos, mayusculas y espacios.
    Si el desplegable esta recargandose por AJAX, espera y vuelve a intentar.
    """
    if pd.isna(texto_buscar) or str(texto_buscar).strip() == "":
        return
    texto_buscar = str(texto_buscar).strip()

    page.wait_for_selector(selector, timeout=5000)
    select_element = page.locator(selector).first

    for _ in range(intentos):
        val_seleccionado = select_element.evaluate(
            r"""(select, texto) => {
                const limpiar = (str) => str.toLowerCase().normalize("NFD").replace(/[\u0300-\u036f]/g, "").trim();
                const txtBuscado = limpiar(texto);

                if (select.options.length <= 1 && select.options[0] && (select.options[0].text.toLowerCase().includes("selecciona") || select.options[0].text.trim() === "")) {
                    return "ESPERAR_AJAX";
                }

                for (let option of select.options) {
                    const txtOption = limpiar(option.text);
                    if (txtOption === txtBuscado || txtOption.includes(txtBuscado) || txtBuscado.includes(txtOption)) {
                        select.value = option.value;
                        select.dispatchEvent(new Event('change', { bubbles: true }));
                        select.dispatchEvent(new Event('input', { bubbles: true }));
                        return option.text;
                    }
                }
                return null;
            }""",
            texto_buscar,
        )

        if val_seleccionado == "ESPERAR_AJAX":
            time.sleep(1.5)
            continue
        if val_seleccionado:
            print(f"    -> Seleccionado con exito: '{val_seleccionado}'")
            return
        time.sleep(1.5)

    print(f"    Alerta: No se encontro coincidencia para '{texto_buscar}' en [{selector}]")


def _procesar_grupo(page, row_grupo, df_integrantes):
    nombre_grupo = str(row_grupo["Grupo / Nombre*"]).strip()
    print(f"\n>> PROCESANDO GRUPO: {nombre_grupo}")

    page.goto(TARGET_URL, wait_until="domcontentloaded")
    page.wait_for_load_state("networkidle")

    page.locator("table thead th:last-child, .bi-plus-circle").filter(visible=True).first.click(force=True)
    page.wait_for_selector("#grupo_estudio_cab_grupo", timeout=5000)

    page.locator("#grupo_estudio_cab_grupo").fill(nombre_grupo)
    seleccionar_opcion_flexible(page, "#grupo_estudio_cab_id_empresa", row_grupo["Empresa*"])
    time.sleep(1.5)
    seleccionar_opcion_flexible(page, "div.col.mb-3:has-text('Sucursal') select, #grupo_estudio_cab_id_sucursal", row_grupo["Sucursal*"])
    seleccionar_opcion_flexible(page, "div.col.mb-3:has-text('Capellan') select, #grupo_estudio_cab_id_usuario", row_grupo["Capellán*"])
    seleccionar_opcion_flexible(page, "div.col.mb-3:has-text('Material') select", row_grupo["Material*"])
    seleccionar_opcion_flexible(page, "div.col.mb-3:has-text('Estado Del Grupo') select", row_grupo["Estado Del Grupo*"])

    page.click("#btn_guardar_grupo_estudio_cab")
    page.wait_for_selector("#grupo_estudio_cab_grupo", state="hidden", timeout=5000)
    page.wait_for_load_state("networkidle")
    time.sleep(2.5)

    integrantes_grupo = df_integrantes[
        df_integrantes["Nombre del Grupo (Exacto)*"].astype(str).str.strip() == nombre_grupo
    ]
    if not integrantes_grupo.empty:
        fila_tabla = page.locator(f"tr:has-text('{nombre_grupo}')").first
        boton_personita = fila_tabla.locator(".bi-people-fill, a:has(.bi-people-fill)").first
        href_relativo = boton_personita.get_attribute("href")
        page.goto(urljoin(page.url, href_relativo))
        page.wait_for_load_state("load")
        page.wait_for_load_state("networkidle")
        time.sleep(2)

        for _, row_integ in integrantes_grupo.iterrows():
            nombre_persona = str(row_integ["Nombre*"]).strip()
            print(f"  -> Anadiendo integrante: {nombre_persona}")
            page.locator(".bi-plus-circle, i.bi-plus-circle").filter(visible=True).first.click(force=True)
            page.wait_for_selector(".modal-body input[name*='nombre'], .modal-body input[id*='nombre']", timeout=5000)

            page.locator(".modal-body input[name*='nombre'], .modal-body input[id*='nombre']").first.fill(nombre_persona)
            if pd.notna(row_integ["Apellido"]):
                page.locator(".modal-body input[name*='apellido'], .modal-body input[id*='apellido']").first.fill(str(row_integ["Apellido"]).strip())
            if pd.notna(row_integ["CI"]):
                page.locator(".modal-body input[name*='ci'], .modal-body input[id*='ci']").first.fill(str(row_integ["CI"]).strip())
            if pd.notna(row_integ["Email"]):
                page.locator(".modal-body input[name*='email'], .modal-body input[id*='email']").first.fill(str(row_integ["Email"]).strip())
            if pd.notna(row_integ["Sección"]):
                seleccionar_opcion_flexible(page, ".modal-body select[name*='seccion'], .modal-body select[id*='seccion']", row_integ["Sección"])

            page.locator(".modal-footer button:has-text('Guardar'), .modal-content .btn-primary").filter(visible=True).first.click(force=True)
            page.wait_for_selector(".modal-body", state="hidden", timeout=5000)
            time.sleep(1.5)

    page.locator("text=Reuniones").filter(visible=True).first.click()
    page.wait_for_load_state("networkidle")
    time.sleep(2)
    page.locator("#btn_show_generar_reuniones_grupo_estudio_cab").click()
    page.wait_for_load_state("networkidle")
    time.sleep(4)


def procesar_carga_real(excel_file=EXCEL_FILE):
    credentials = validar_credenciales()
    inicializar_estado_carga(excel_file)
    df_grupos = leer_grupos(excel_file)
    df_integrantes = leer_integrantes(excel_file)
    pendientes = df_grupos[df_grupos[STATUS_COL].astype(str).str.strip().str.upper() == "PENDIENTE"]

    print(f"Detectados {len(pendientes)} grupos pendientes.")
    if pendientes.empty:
        return {"procesados": 0, "errores": 0}

    procesados = 0
    errores = 0
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True, args=CHROMIUM_ARGS)
        page = browser.new_page()

        page.goto(WP_LOGIN_URL)
        page.fill("#user_login", credentials["user"])
        page.fill("#user_pass", credentials["password"])
        page.click("#wp-submit")
        page.wait_for_selector("#adminmenu, #wpadminbar", timeout=15000)
        page.wait_for_timeout(2000)

        for index, row_grupo in pendientes.iterrows():
            excel_row = int(index) + 2
            try:
                _procesar_grupo(page, row_grupo, df_integrantes)
                actualizar_estado_fila(excel_row, "PROCESADO", green_fill, excel_file)
                procesados += 1
            except Exception as exc:
                print(f"ERROR en fila {excel_row}: {exc}")
                actualizar_estado_fila(excel_row, "ERROR", red_fill, excel_file)
                errores += 1

        browser.close()

    return {"procesados": procesados, "errores": errores}


if __name__ == "__main__":
    procesar_carga_real()
