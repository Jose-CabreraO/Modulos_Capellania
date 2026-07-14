import os
import re
import tempfile
import textwrap
import unicodedata
from io import BytesIO

os.environ.setdefault("MPLCONFIGDIR", tempfile.gettempdir())
import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


DAY_ORDER = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado"]
DAY_ALIASES = {
    "lunes": "Lunes",
    "martes": "Martes",
    "miercoles": "Miércoles",
    "miércoles": "Miércoles",
    "jueves": "Jueves",
    "viernes": "Viernes",
    "sabado": "Sábado",
    "sábado": "Sábado",
}
EXPECTED_COLUMNS = [
    "ListadoId",
    "Capellan",
    "Empresa",
    "Sucursal",
    "Sección",
    "Dia",
    "Hora",
    "RangoTiempoCarga",
    "FechaAct",
]


def normalize_text(value):
    text = "" if value is None else str(value)
    text = unicodedata.normalize("NFD", text.lower())
    text = "".join(char for char in text if unicodedata.category(char) != "Mn")
    return re.sub(r"\s+", " ", text).strip()


def normalize_day(value):
    key = normalize_text(value)
    return DAY_ALIASES.get(key, str(value).strip())


def _is_day_token(value):
    return normalize_text(value) in {normalize_text(day) for day in DAY_ORDER}


def _read_by_wide_separator(text):
    rows = [line.strip() for line in text.splitlines() if line.strip()]
    if not rows:
        raise ValueError("No se detectaron filas para procesar.")

    header = re.split(r"\s{2,}|\t+", rows[0].strip())
    if len(header) < 5:
        return None

    parsed_rows = [re.split(r"\s{2,}|\t+", row.strip()) for row in rows[1:]]
    valid_rows = [row for row in parsed_rows if len(row) == len(header)]
    if not valid_rows:
        return None

    return pd.DataFrame(valid_rows, columns=header)


def _parse_line_fallback(line):
    tokens = line.split()
    if len(tokens) < 9 or not tokens[0].isdigit():
        return None

    day_idx = next((idx for idx, token in enumerate(tokens) if _is_day_token(token)), None)
    if day_idx is None or day_idx + 4 >= len(tokens):
        return None

    listado_id = tokens[0]
    dia = tokens[day_idx]
    hora = tokens[day_idx + 1]
    rango = " ".join(tokens[day_idx + 2 : day_idx + 4])
    fecha = tokens[day_idx + 4]

    middle = tokens[1:day_idx]
    if len(middle) < 4:
        return None

    seccion = middle[-1]
    sucursal = middle[-2]
    capellan_tokens = middle[:4] if len(middle) >= 6 else middle[: max(1, len(middle) - 3)]
    empresa_tokens = middle[len(capellan_tokens) : -2]

    return {
        "ListadoId": listado_id,
        "Capellan": " ".join(capellan_tokens),
        "Empresa": " ".join(empresa_tokens) or "---",
        "Sucursal": sucursal,
        "Sección": seccion,
        "Dia": dia,
        "Hora": hora,
        "RangoTiempoCarga": rango,
        "FechaAct": fecha,
    }


def parse_schedule_text(text):
    if not text or not text.strip():
        raise ValueError("Pegá el listado de horarios antes de procesar.")

    df = _read_by_wide_separator(text)
    if df is None:
        lines = [line.strip() for line in text.splitlines() if line.strip()]
        data_lines = lines[1:] if lines and "ListadoId" in lines[0] else lines
        records = [_parse_line_fallback(line) for line in data_lines]
        records = [record for record in records if record]
        if not records:
            raise ValueError("No se pudo interpretar el texto pegado. Verificá encabezados y columnas.")
        df = pd.DataFrame(records)

    rename_map = {}
    normalized_columns = {normalize_text(col): col for col in df.columns}
    for expected in EXPECTED_COLUMNS:
        found = normalized_columns.get(normalize_text(expected))
        if found:
            rename_map[found] = expected
    df = df.rename(columns=rename_map)

    missing = [col for col in ["Empresa", "Sucursal", "Dia", "Hora"] if col not in df.columns]
    if missing:
        raise ValueError("Faltan columnas obligatorias: " + ", ".join(missing))

    df = df.dropna(subset=["Dia", "Hora"]).copy()
    df["Dia"] = df["Dia"].map(normalize_day)
    df = df[df["Dia"].isin(DAY_ORDER)]
    if df.empty:
        raise ValueError("No se detectaron filas con días válidos de Lunes a Sábado.")

    parsed_hours = pd.to_datetime(df["Hora"].astype(str), format="%H:%M", errors="coerce")
    df = df[parsed_hours.notna()].copy()
    if df.empty:
        raise ValueError("No se detectaron horas válidas con formato HH:MM.")

    df["HoraDt"] = parsed_hours[parsed_hours.notna()]
    df["Hora"] = df["HoraDt"].dt.strftime("%H:%M")
    df["Dia_Norm"] = df["Dia"]
    df["Asignacion"] = df.apply(_format_assignment, axis=1)
    return df


def _hour_block_label(hour):
    return f"{int(hour):02d}:00"


def _format_assignment(row):
    empresa = str(row["Empresa"]).strip()
    sucursal = str(row["Sucursal"]).strip()
    line_one = textwrap.fill(f"{row['Hora']} - {empresa}", width=18, break_long_words=False)
    line_two = textwrap.fill(f"({sucursal})", width=18, break_long_words=False)
    return f"{line_one}\n{line_two}"


def build_schedule_grid(df):
    work = df.copy()
    if "Dia_Norm" not in work.columns:
        work["Dia_Norm"] = work["Dia"].map(normalize_day)
    if "Asignacion" not in work.columns:
        work["Asignacion"] = work.apply(_format_assignment, axis=1)

    min_hour = int(work["HoraDt"].dt.hour.min())
    max_hour = int(work["HoraDt"].dt.hour.max())
    hours = [_hour_block_label(hour) for hour in range(min_hour, max_hour + 1)]

    work["Hora"] = work["HoraDt"].dt.hour.map(_hour_block_label)
    grouped = (
        work.sort_values(["HoraDt", "Empresa", "Sucursal"])
        .groupby(["Hora", "Dia_Norm"], as_index=False)["Asignacion"]
        .agg(lambda values: "\n\n".join(values))
    )

    cuadrilla = grouped.pivot(index="Hora", columns="Dia_Norm", values="Asignacion")
    cuadrilla = cuadrilla.reindex(index=hours, columns=DAY_ORDER).fillna("—")
    return cuadrilla.reset_index()


def _clean_capellan_name(row):
    capellan = str(row.get("Capellan", "")).strip()
    empresa = str(row.get("Empresa", "")).strip()
    if not capellan:
        return ""

    if empresa:
        capellan_norm = normalize_text(capellan)
        empresa_norm = normalize_text(empresa)
        if capellan_norm.endswith(empresa_norm):
            capellan = capellan[: -len(empresa)].strip()
        elif capellan_norm.startswith(empresa_norm):
            capellan = capellan[len(empresa) :].strip()

    return re.sub(r"\s+", " ", capellan).strip()


def get_capellan_title(df):
    if "Capellan" not in df.columns:
        return "Horario"

    capellanes = [_clean_capellan_name(row) for _, row in df.iterrows()]
    capellanes = pd.Series(capellanes, dtype="object")
    capellanes = capellanes[capellanes != ""]
    if capellanes.empty:
        return "Horario"
    return f"Horario - {capellanes.iloc[0]}"


def highlight_occupied_cells(value):
    return "background-color: #FFFFFF; color: #111827; font-weight: 500;"


def _wrapped_cell(text, width=18):
    if text == "—":
        return text
    parts = str(text).splitlines()
    return "\n".join(textwrap.fill(part, width=width, break_long_words=False) for part in parts)


def generate_schedule_pdf(grid, title="Organizador de horarios"):
    fig, ax = plt.subplots(figsize=(11.69, 8.27))
    fig.patch.set_facecolor("#FFFFFF")
    ax.axis("off")
    fig.text(0.02, 0.955, title, fontsize=18, fontweight="bold", color="#1F2937", ha="left", va="top")

    columns = list(grid.columns)
    data_width = 15 if len(grid) > 8 else 18
    cell_text = [
        [_wrapped_cell(row[column], width=data_width if column != "Hora" else 8) for column in columns]
        for _, row in grid.iterrows()
    ]
    cell_colours = [
        ["#BAE6FD" if column == "Hora" else ("#F8FAFC" if row_idx % 2 else "#FFFFFF") for column in columns]
        for row_idx, _ in enumerate(cell_text)
    ]

    tabla = ax.table(
        cellText=cell_text,
        colLabels=columns,
        cellColours=cell_colours,
        colColours=["#BAE6FD"] * len(columns),
        colWidths=[0.095] + [0.150] * len(DAY_ORDER),
        cellLoc="center",
        loc="center",
        bbox=[0.02, 0.025, 0.96, 0.865],
    )
    tabla.auto_set_font_size(False)
    body_fontsize = 8 if len(grid) > 9 else 9
    header_fontsize = 10

    for (row, col), cell in tabla.get_celld().items():
        cell.set_edgecolor("#94A3B8")
        cell.set_linewidth(0.55)
        cell.PAD = 0.08
        cell.get_text().set_ha("center")
        cell.get_text().set_va("center")
        if row == 0:
            cell.set_facecolor("#BAE6FD")
            cell.set_text_props(weight="bold", color="#111827", fontsize=header_fontsize)
        elif col == 0:
            cell.set_facecolor("#BAE6FD")
            cell.set_text_props(weight="bold", color="#111827", fontsize=header_fontsize)
        else:
            cell.set_text_props(color="#111827", fontsize=body_fontsize)

    line_counts = [max(str(value).count("\n") + 1 for value in row_values) for row_values in cell_text]
    total_units = 1.25 + sum(max(1.35, count * 0.86) for count in line_counts)
    alto_base = 0.865 / total_units

    for col_idx in range(len(columns)):
        tabla[(0, col_idx)].set_height(alto_base * 1.25)

    for row_idx, cantidad_de_lineas in enumerate(line_counts, start=1):
        row_height = alto_base * max(1.35, cantidad_de_lineas * 0.86)
        for col_idx in range(len(columns)):
            tabla[(row_idx, col_idx)].set_height(row_height)

    fig.subplots_adjust(left=0.0, right=1.0, top=0.91, bottom=0.0)
    output = BytesIO()
    fig.savefig(output, format="pdf", orientation="landscape")
    plt.close(fig)
    output.seek(0)
    return output


def generate_schedule_excel(grid, title="Organizador de horarios"):
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        grid.to_excel(writer, index=False, startrow=2, sheet_name="Horario")
        worksheet = writer.sheets["Horario"]

        max_col = len(grid.columns)
        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
        title_cell = worksheet.cell(row=1, column=1)
        title_cell.value = title
        title_cell.font = Font(bold=True, size=16, color="1F2937")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        header_fill = PatternFill("solid", fgColor="BAE6FD")
        hour_fill = PatternFill("solid", fgColor="E0F2FE")
        thin_side = Side(style="thin", color="CBD5E1")
        border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

        for row in worksheet.iter_rows(min_row=3, max_row=worksheet.max_row, min_col=1, max_col=max_col):
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = border
                if cell.row == 3:
                    cell.fill = header_fill
                    cell.font = Font(bold=True, color="111827")
                elif cell.column == 1:
                    cell.fill = hour_fill
                    cell.font = Font(bold=True, color="111827")

        worksheet.column_dimensions["A"].width = 12
        for col_idx in range(2, max_col + 1):
            worksheet.column_dimensions[get_column_letter(col_idx)].width = 24
        for row_idx in range(4, worksheet.max_row + 1):
            max_lines = max(
                str(worksheet.cell(row=row_idx, column=col_idx).value or "").count("\n") + 1
                for col_idx in range(1, max_col + 1)
            )
            worksheet.row_dimensions[row_idx].height = max(28, min(95, max_lines * 15))

        worksheet.freeze_panes = "B4"
        worksheet.sheet_view.showGridLines = False

    output.seek(0)
    return output


def build_ui():
    import streamlit as st

    st.title("Organizador de Horarios")
    raw_text = st.text_area(
        "Pegá el listado copiado desde la web",
        height=220,
        placeholder=(
            "ListadoId Capellan Empresa Sucursal Sección Dia Hora RangoTiempoCarga FechaAct\n"
            "1878 Souberlich Vidal Rodrigo Sebastin Inverfín S.A.E.C.A. Areguá General Martes 09:30 1 Hora/s 2026-05-12"
        ),
    )

    if not raw_text.strip():
        return

    try:
        parsed = parse_schedule_text(raw_text)
        grid = build_schedule_grid(parsed)
        title = get_capellan_title(parsed)
    except Exception as exc:
        st.error(f"No se pudo procesar el texto pegado: {exc}")
        return

    st.subheader(title)
    st.caption(f"Registros procesados: {len(parsed)}")
    styled = (
        grid.style
        .map(highlight_occupied_cells, subset=DAY_ORDER)
        .set_properties(subset=["Hora"], **{"background-color": "#BAE6FD", "font-weight": "700", "color": "#075985"})
        .set_table_styles([
            {"selector": "th", "props": [("background-color", "#BAE6FD"), ("color", "#1E3A8A"), ("font-weight", "700")]},
        ])
    )
    st.dataframe(styled, width="stretch", hide_index=True)

    pdf = generate_schedule_pdf(grid, title=title)
    excel = generate_schedule_excel(grid, title=title)
    col_pdf, col_excel = st.columns(2)
    with col_pdf:
        st.download_button(
            "Descargar PDF para imprimir",
            data=pdf,
            file_name="organizador_horarios.pdf",
            mime="application/pdf",
            width="stretch",
        )
    with col_excel:
        st.download_button(
            "Descargar Excel",
            data=excel,
            file_name="organizador_horarios.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            width="stretch",
        )
