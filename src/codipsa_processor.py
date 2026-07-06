import calendar
import difflib
import re
import unicodedata
from copy import copy
from datetime import datetime
from io import BytesIO

import pandas as pd
from openpyxl import load_workbook


SHEET_BASE = "CANT-PROD"
SHEET_SEMANAS = "Semanas"
NAME_COLUMN = "FAMILIAS"
VISIT_COUNT_COLUMN = "CANT VECES VISITADO"

CONTROL_SHEETS = [
    "Control",
    "Cargas realizadas",
    "No encontradas",
    "Duplicados",
    "Errores de fecha",
    "Coincidencias revisables",
]

EXPECTED_COLUMNS = {
    "Codipsa 1": [
        "SEMANAS",
        "Nombre Completo del Colaborador",
        "Cantidad de participantes",
        "Tema de la reflexión",
        "Dia de la visita",
        "Decisiones por Cristo",
        "Validación",
    ],
    "Codipsa 3": [
        "SEMANAS",
        "Nombre Completo del Colaborador",
        "Cantidad de participantes",
        "Tema de la reflexión",
        "Dia de la visita",
        "Barrio o zona de la visita",
        "Decisiones por Cristo",
        "Validación",
    ],
}


def normalize_name(value):
    text = "" if value is None else str(value)
    text = unicodedata.normalize("NFD", text.upper())
    text = "".join(char for char in text if unicodedata.category(char) != "Mn")
    text = re.sub(r"[^A-ZÑ0-9\s]", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def normalize_header(value):
    return normalize_name(value).replace("Ñ", "N")


def get_iso_weeks_for_month(year, month):
    _, last_day = calendar.monthrange(year, month)
    weeks = []
    for day in range(1, last_day + 1):
        week = datetime(year, month, day).date().isocalendar().week
        if week not in weeks:
            weeks.append(week)
    return weeks


def load_semana_sheet(uploaded_file, group):
    try:
        df = pd.read_excel(uploaded_file, sheet_name=SHEET_SEMANAS)
    except ValueError as exc:
        raise ValueError(f"El archivo de Google Sheets no contiene la hoja '{SHEET_SEMANAS}'.") from exc

    df.columns = [str(col).strip() for col in df.columns]
    expected = EXPECTED_COLUMNS[group]
    normalized_columns = {normalize_header(col): col for col in df.columns}
    normalized_keys = list(normalized_columns.keys())
    rename_map = {}
    missing = []
    for col in expected:
        expected_key = normalize_header(col)
        found = normalized_columns.get(expected_key)
        if not found:
            close = difflib.get_close_matches(expected_key, normalized_keys, n=1, cutoff=0.82)
            found = normalized_columns.get(close[0]) if close else None
        if found:
            rename_map[found] = col
        else:
            missing.append(col)
    if missing:
        raise ValueError("Faltan columnas en Semanas: " + ", ".join(missing))

    df = df.rename(columns=rename_map)
    return df.dropna(how="all")


def find_header_row(ws, required_column=NAME_COLUMN, max_scan_rows=25):
    required = normalize_header(required_column)
    for row in range(1, min(ws.max_row, max_scan_rows) + 1):
        headers = [normalize_header(ws.cell(row=row, column=col).value) for col in range(1, ws.max_column + 1)]
        if required in headers:
            return row
    raise ValueError(f"No se encontro la columna '{required_column}' en la hoja '{SHEET_BASE}'.")


def header_map(ws, header_row):
    return {
        normalize_header(ws.cell(row=header_row, column=col).value): col
        for col in range(1, ws.max_column + 1)
        if ws.cell(row=header_row, column=col).value not in (None, "")
    }


def detect_month_columns(ws, header_row, month, year):
    weeks = get_iso_weeks_for_month(year, month)
    month_names = {
        normalize_name(calendar.month_name[month]),
        normalize_name(calendar.month_abbr[month]),
    }
    month_names.update({
        normalize_name(name)
        for name in [
            "ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO",
            "JULIO", "AGOSTO", "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE",
        ][month - 1:month]
    })

    candidates = []
    for col in range(1, ws.max_column + 1):
        texts = []
        for row in range(max(1, header_row - 3), header_row + 2):
            value = ws.cell(row=row, column=col).value
            if value not in (None, ""):
                texts.append(normalize_name(value))
        combined = " ".join(texts)
        if any(month_name and month_name in combined for month_name in month_names):
            candidates.append(col)

    if len(candidates) < len(weeks):
        fallback = []
        for col in range(1, ws.max_column + 1):
            combined = " ".join(
                normalize_name(ws.cell(row=row, column=col).value)
                for row in range(max(1, header_row - 3), header_row + 2)
                if ws.cell(row=row, column=col).value not in (None, "")
            )
            if any(str(week) in combined for week in weeks):
                fallback.append(col)
        candidates = sorted(set(candidates + fallback))

    candidates = sorted(candidates)
    if len(candidates) < len(weeks):
        raise ValueError(
            f"No se pudieron detectar suficientes columnas para el mes seleccionado. "
            f"Semanas ISO esperadas: {weeks}. Columnas detectadas: {len(candidates)}."
        )

    return dict(zip(weeks, candidates[: len(weeks)]))


def build_people_index(ws, header_row):
    headers = header_map(ws, header_row)
    name_col = headers.get(normalize_header(NAME_COLUMN))
    if not name_col:
        raise ValueError(f"No se encontro la columna '{NAME_COLUMN}' en '{SHEET_BASE}'.")

    people = []
    for row in range(header_row + 1, ws.max_row + 1):
        name = ws.cell(row=row, column=name_col).value
        normalized = normalize_name(name)
        if normalized:
            people.append({"row": row, "name": name, "normalized": normalized})
    return people, headers


def match_person(raw_name, people):
    normalized = normalize_name(raw_name)
    if not normalized:
        return None, "no_encontrada", []

    exact = [person for person in people if person["normalized"] == normalized]
    if len(exact) == 1:
        return exact[0], "exacta", exact
    if len(exact) > 1:
        return None, "revisable", exact

    partial = [
        person for person in people
        if normalized in person["normalized"] or person["normalized"] in normalized
    ]
    if len(partial) == 1:
        return partial[0], "parcial", partial
    if len(partial) > 1:
        return None, "revisable", partial

    names = [person["normalized"] for person in people]
    similar_names = difflib.get_close_matches(normalized, names, n=3, cutoff=0.86)
    similar = [person for person in people if person["normalized"] in similar_names]
    if len(similar) == 1:
        return similar[0], "similitud", similar
    if len(similar) > 1:
        return None, "revisable", similar

    return None, "no_encontrada", []


def parse_visit_date(value):
    if pd.isna(value):
        return None
    if isinstance(value, datetime):
        return value.date()
    parsed = pd.to_datetime(value, dayfirst=True, errors="coerce")
    if pd.isna(parsed):
        return None
    return parsed.date()


def clone_first_data_row_style(ws, header_row, target_row):
    template_row = header_row + 1
    if template_row >= target_row:
        return
    for col in range(1, ws.max_column + 1):
        source = ws.cell(row=template_row, column=col)
        target = ws.cell(row=target_row, column=col)
        if source.has_style:
            target._style = copy(source._style)
        if source.number_format:
            target.number_format = source.number_format


def append_control_record(collection, **kwargs):
    collection.append(kwargs)


def process_visits(base_file, semanas_file, group, month, year):
    wb = load_workbook(base_file)
    if SHEET_BASE not in wb.sheetnames:
        raise ValueError(f"El archivo base no contiene la hoja '{SHEET_BASE}'.")
    ws = wb[SHEET_BASE]

    df_sem = load_semana_sheet(semanas_file, group)
    header_row = find_header_row(ws)
    people, headers = build_people_index(ws, header_row)
    week_columns = detect_month_columns(ws, header_row, month, year)
    count_col = headers.get(normalize_header(VISIT_COUNT_COLUMN))

    stats = {
        "total_registros": int(len(df_sem)),
        "registros_validos": 0,
        "cargas_realizadas": 0,
        "no_encontradas": 0,
        "duplicados": 0,
        "errores_fecha": 0,
        "coincidencias_exactas": 0,
        "coincidencias_parciales": 0,
        "coincidencias_similitud": 0,
    }
    controls = {sheet: [] for sheet in CONTROL_SHEETS if sheet != "Control"}
    loaded_slots = set()

    for idx, record in df_sem.iterrows():
        row_num = int(idx) + 2
        raw_name = record.get("Nombre Completo del Colaborador")
        visit_date = parse_visit_date(record.get("Dia de la visita"))

        if not visit_date or visit_date.month != month or visit_date.year != year:
            stats["errores_fecha"] += 1
            append_control_record(
                controls["Errores de fecha"],
                fila=row_num,
                nombre=raw_name,
                dia_de_la_visita=record.get("Dia de la visita"),
                motivo="Fecha vacia, invalida o fuera del mes seleccionado",
            )
            continue

        week = visit_date.isocalendar().week
        target_col = week_columns.get(week)
        if not target_col:
            stats["errores_fecha"] += 1
            append_control_record(
                controls["Errores de fecha"],
                fila=row_num,
                nombre=raw_name,
                dia_de_la_visita=str(visit_date),
                motivo=f"No hay columna detectada para semana ISO {week}",
            )
            continue

        person, match_type, candidates = match_person(raw_name, people)
        if match_type == "exacta":
            stats["coincidencias_exactas"] += 1
        elif match_type == "parcial":
            stats["coincidencias_parciales"] += 1
        elif match_type == "similitud":
            stats["coincidencias_similitud"] += 1

        if match_type == "no_encontrada":
            stats["no_encontradas"] += 1
            append_control_record(
                controls["No encontradas"],
                fila=row_num,
                nombre=raw_name,
                nombre_normalizado=normalize_name(raw_name),
                dia_de_la_visita=str(visit_date),
            )
            continue

        if match_type == "revisable":
            append_control_record(
                controls["Coincidencias revisables"],
                fila=row_num,
                nombre=raw_name,
                nombre_normalizado=normalize_name(raw_name),
                candidatos=" | ".join(str(candidate["name"]) for candidate in candidates),
                dia_de_la_visita=str(visit_date),
            )
            continue

        stats["registros_validos"] += 1
        slot = (person["row"], week)
        if slot in loaded_slots or ws.cell(row=person["row"], column=target_col).value not in (None, ""):
            stats["duplicados"] += 1
            append_control_record(
                controls["Duplicados"],
                fila=row_num,
                nombre=raw_name,
                familia_encontrada=person["name"],
                semana_iso=week,
                dia_de_la_visita=str(visit_date),
            )
            continue

        ws.cell(row=person["row"], column=target_col, value=visit_date.day)
        loaded_slots.add(slot)
        stats["cargas_realizadas"] += 1
        append_control_record(
            controls["Cargas realizadas"],
            fila=row_num,
            nombre=raw_name,
            familia_encontrada=person["name"],
            tipo_coincidencia=match_type,
            semana_iso=week,
            dia_cargado=visit_date.day,
        )

        if count_col:
            current = ws.cell(row=person["row"], column=count_col).value
            try:
                current = int(current or 0)
            except (TypeError, ValueError):
                current = 0
            ws.cell(row=person["row"], column=count_col, value=current + 1)

    write_control_sheets(wb, stats, controls, group, month, year)
    output = BytesIO()
    wb.save(output)
    wb.close()
    output.seek(0)
    return output, stats


def write_dataframe_sheet(wb, sheet_name, records):
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    if not records:
        ws.append(["Sin registros"])
        return

    df = pd.DataFrame(records)
    ws.append(list(df.columns))
    for row in df.itertuples(index=False, name=None):
        ws.append(list(row))


def write_control_sheets(wb, stats, controls, group, month, year):
    if "Control" in wb.sheetnames:
        del wb["Control"]
    ws = wb.create_sheet("Control")
    ws.append(["Campo", "Valor"])
    ws.append(["Grupo", group])
    ws.append(["Mes", month])
    ws.append(["Año", year])
    for key, value in stats.items():
        ws.append([key, value])

    for sheet_name, records in controls.items():
        write_dataframe_sheet(wb, sheet_name, records)


def build_ui():
    import streamlit as st

    st.title("Procesador Codipsa")

    col1, col2, col3 = st.columns(3)
    group = col1.selectbox("Grupo", ["Codipsa 1", "Codipsa 3"])
    month = col2.selectbox("Mes de trabajo", list(range(1, 13)), index=datetime.today().month - 1)
    year = col3.number_input("Año de trabajo", min_value=2020, max_value=2100, value=datetime.today().year, step=1)

    base_file = st.file_uploader("Archivo Excel base con hoja CANT-PROD", type=["xlsx"])
    semanas_file = st.file_uploader("Archivo Excel exportado desde Google Sheets con hoja Semanas", type=["xlsx"])

    if not base_file or not semanas_file:
        return

    if st.button("Procesar visitas", type="primary"):
        try:
            output, stats = process_visits(base_file, semanas_file, group, int(month), int(year))
        except ValueError as exc:
            st.error(str(exc))
            return
        except Exception as exc:
            st.error(f"No se pudo procesar el archivo: {exc}")
            return

        st.subheader("Resumen")
        cols = st.columns(4)
        labels = [
            ("Total Semanas", "total_registros"),
            ("Registros validos", "registros_validos"),
            ("Cargas realizadas", "cargas_realizadas"),
            ("No encontradas", "no_encontradas"),
            ("Duplicados", "duplicados"),
            ("Errores de fecha", "errores_fecha"),
            ("Exactas", "coincidencias_exactas"),
            ("Parciales", "coincidencias_parciales"),
            ("Similitud", "coincidencias_similitud"),
        ]
        for idx, (label, key) in enumerate(labels):
            cols[idx % 4].metric(label, stats[key])

        filename = f"codipsa_{group.lower().replace(' ', '_')}_{year}_{int(month):02d}_procesado.xlsx"
        st.download_button(
            "Descargar Excel procesado",
            data=output,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
